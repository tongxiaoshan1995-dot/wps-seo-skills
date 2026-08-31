#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WPS SEO 每日关键词规划脚本
==========================
基于《SEO关键词0828.xlsx》四类策略词（品牌词 / 竞品词 / 通用词 / 功能词），
按「月均搜索量」降序提炼每日关键词，按既定配比分配词量，
输出「策略组别 / 今日关键词 / 搜索意图」四列表，并自动跨日去重：
每次调用给出的关键词与历史已用词不重复。

默认配比（可在 RATIO 常量调整，务必合计 100%）：
  品牌词 35% / 竞品词 10% / 通用词 15% / 功能词 40%

用法示例：
    python plan_daily.py
    python plan_daily.py --total 20
    python plan_daily.py --day 3 --out /path/第3天规划.md
    python plan_daily.py --xlsx /path/SEO关键词0828.xlsx
    python plan_daily.py --csv /custom/path/wps_daily_kw.csv --state /custom/path/state.json
    python plan_daily.py --reset   # 清空去重历史，从第 1 天重新开始
"""
import argparse
import glob
import json
import os
import sys
from datetime import datetime

# 四类策略词配比（合计需为 100；顺序即输出顺序）
RATIO = [
    ("品牌词", 0.35),
    ("竞品词", 0.10),
    ("通用词", 0.15),
    ("功能词", 0.40),
]

CATEGORIES = [name for name, _ in RATIO]

# 每类词的搜索意图标签（落到落地页承接逻辑）
INTENT = {
    "品牌词": "品牌/官网/正版入口，路径短、转化快，重承接与信任，承接品牌直达流量",
    "竞品词": "竞品/对比/替代需求，适合客观对比页与FAQ，拦截竞品搜索流量",
    "通用词": "办公软件泛需求，适合知识科普与工具页，拓展新用户认知",
    "功能词": "word/excel/ppt等具体功能需求，意图精准，适合功能教程页与工具页",
}

# 频率列：月均搜索量（Excel 中“月均搜索量”列）
FREQ_COL = "月均搜索量"


# ---------------------------------------------------------------------------
# 数据源定位与解析
# ---------------------------------------------------------------------------
def find_workspace_xlsx():
    """按优先级定位关键词 Excel：当前/最近工作区 > 下载目录"""
    home = os.path.expanduser("~")
    bases = []
    ws = os.environ.get("WORKSPACE_DIR")
    if ws:
        bases.append(ws)
    bases.append(os.path.join(home, "Documents", "lingxi-claw"))
    bases.append(os.path.join(home, "Downloads"))
    for base in bases:
        if os.path.isdir(base):
            found = sorted(glob.glob(os.path.join(base, "SEO关键词*.xlsx")),
                           key=os.path.getmtime, reverse=True)
            if found:
                return found[0]
    return None


def find_builtin_csv():
    skill_data = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "assets", "data", "wps_daily_kw.csv")
    return skill_data if os.path.isfile(skill_data) else None


def load_from_xlsx(xlsx_path):
    """从 Excel 四类词 sheet 读取，展平多列块结构为统一 DataFrame。"""
    import pandas as pd

    def parse_block(sheet_name, category):
        raw = pd.read_excel(xlsx_path, sheet_name=sheet_name)
        blocks = []
        for c in raw.columns:
            if "关键词" not in str(c):
                continue
            suffix = str(c).replace("关键词", "")
            kw = raw[[c]].copy()
            kw.columns = ["关键词"]
            feat, vol, diff = f"特色{suffix}", f"月均搜索量{suffix}", f"竞争激烈程度{suffix}"
            if feat in raw.columns:
                kw["特色"] = raw[feat]
            if vol in raw.columns:
                kw["月均搜索量"] = pd.to_numeric(raw[vol], errors="coerce")
            if diff in raw.columns:
                kw["竞争激烈程度"] = raw[diff]
            kw["类别"] = category
            blocks.append(kw)
        out = pd.concat(blocks, ignore_index=True)
        out = out.dropna(subset=["关键词"])
        out["关键词"] = out["关键词"].astype(str).str.strip()
        out = out[out["关键词"].str.lower() != "nan"]
        out = out.drop_duplicates(subset=["关键词"])
        for col in ("月均搜索量", "特色", "竞争激烈程度"):
            if col not in out.columns:
                out[col] = None
        return out[["类别", "关键词", "月均搜索量", "特色", "竞争激烈程度"]]

    sheet_map = {"品牌词": "wps品牌词", "竞品词": "竞品词", "通用词": "通用词", "功能词": "功能词"}
    frames = [parse_block(sheet_map[c], c) for c in CATEGORIES]
    df = pd.concat(frames, ignore_index=True)
    df["月均搜索量"] = pd.to_numeric(df["月均搜索量"], errors="coerce")
    df = df.dropna(subset=["关键词"]).drop_duplicates(subset=["关键词"], keep="first")
    return df


def load_from_csv(csv_path):
    import pandas as pd
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    df["月均搜索量"] = pd.to_numeric(df["月均搜索量"], errors="coerce")
    df = df.dropna(subset=["关键词"]).drop_duplicates(subset=["关键词"], keep="first")
    df["类别"] = df["类别"].astype(str).str.strip()
    df = df[df["类别"].isin(CATEGORIES)]
    return df


# ---------------------------------------------------------------------------
# 配比取整（最大余数法，保证合计 = 总数）
# ---------------------------------------------------------------------------
def allocate_quota(total):
    """按配比把 total 分配给各策略组，返回 {类别: 词数}，合计严格等于 total。"""
    raw = [(name, ratio * total) for name, ratio in RATIO]
    floors = [(name, int(q)) for name, q in raw]
    base = sum(n for _, n in floors)
    remain = total - base
    # 按小数部分降序分配剩余名额
    order = sorted(raw, key=lambda x: x[1] - int(x[1]), reverse=True)
    result = {name: n for name, n in floors}
    for name, _ in order[:remain]:
        result[name] += 1
    return result


# ---------------------------------------------------------------------------
# 状态读写（按组记录已用词，轮转时只清空该组）
# ---------------------------------------------------------------------------
def find_state_path(explicit=None):
    if explicit:
        return explicit
    home = os.path.expanduser("~")
    base = os.path.join(home, "Documents", "lingxi-claw")
    if os.path.isdir(base):
        found = sorted(glob.glob(os.path.join(base, "*", ".wps_seo_state.json")),
                       key=os.path.getmtime, reverse=True)
        if found:
            return found[0]
        wss = sorted([d for d in glob.glob(os.path.join(base, "*")) if os.path.isdir(d)],
                     key=os.path.getmtime, reverse=True)
        if wss:
            return os.path.join(wss[0], ".wps_seo_state.json")
    ws = os.environ.get("WORKSPACE_DIR")
    if ws:
        return os.path.join(ws, ".wps_seo_state.json")
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(os.path.join(skill_root, "data"), exist_ok=True)
    return os.path.join(skill_root, "data", "state.json")


def load_state(path):
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"last_day": 0, "used_by_cat": {}, "updated_at": None}


def save_state(path, state):
    state["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def pick_for_category(sub, used_cat, quota):
    """从子集 sub 中剔除 used_cat 后按月均搜索量降序取 quota 个。
    未用词不足时轮转该组：清空本组历史，从组头重新取词（保证每天不重复）。
    返回 (chosen, rotated)。rotated=True 表示该组触发了轮转。"""
    cand = sub[~sub["关键词"].isin(used_cat)].copy()
    if len(cand) >= quota:
        cand = cand.sort_values(FREQ_COL, ascending=False, na_position="last")
        return cand.head(quota)["关键词"].tolist(), False
    # 未用词不足：轮转本组，从组头重取
    cand = sub.sort_values(FREQ_COL, ascending=False, na_position="last")
    return cand.head(quota)["关键词"].tolist(), True


def main():
    ap = argparse.ArgumentParser(description="WPS SEO 每日关键词规划")
    ap.add_argument("--xlsx", help="关键词 Excel 路径（默认自动定位最新 SEO关键词*.xlsx）")
    ap.add_argument("--csv", help="统一词库 CSV 路径（优先级低于 --xlsx）")
    ap.add_argument("--state", help="状态 JSON 路径（默认自动定位）")
    ap.add_argument("--day", type=int, help="指定第几天（默认自动递增）")
    ap.add_argument("--total", type=int, default=30, help="每日关键词总数（默认 30，可调 20~30）")
    ap.add_argument("--out", help="输出文件路径（.md 或 .xlsx）")
    ap.add_argument("--reset", action="store_true", help="清空去重历史，从第 1 天开始")
    args = ap.parse_args()

    if not (20 <= args.total <= 30):
        sys.exit(f"[错误] --total 需在 20~30 之间，当前为 {args.total}")

    import pandas as pd

    # 数据源：显式 xlsx > 显式 csv > 自动定位 xlsx > 内置 csv
    df = None
    src_desc = None
    if args.xlsx and os.path.isfile(args.xlsx):
        df = load_from_xlsx(args.xlsx)
        src_desc = os.path.basename(args.xlsx)
    elif args.csv and os.path.isfile(args.csv):
        df = load_from_csv(args.csv)
        src_desc = os.path.basename(args.csv)
    else:
        auto_xlsx = find_workspace_xlsx()
        if auto_xlsx:
            df = load_from_xlsx(auto_xlsx)
            src_desc = os.path.basename(auto_xlsx)
        else:
            csv_path = find_builtin_csv()
            if csv_path:
                df = load_from_csv(csv_path)
                src_desc = os.path.basename(csv_path)
    if df is None:
        sys.exit("[错误] 找不到关键词数据源，请用 --xlsx 或 --csv 指定路径")
    print(f"[数据] 关键词库：{src_desc}（{len(df)} 词）", file=sys.stderr)

    state_path = find_state_path(args.state)
    if args.reset:
        state = {"last_day": 0, "used_by_cat": {}, "updated_at": None}
    else:
        state = load_state(state_path)
    used_by_cat = state.get("used_by_cat", {})
    day = args.day if args.day else (state.get("last_day", 0) + 1)

    quota = allocate_quota(args.total)
    print(f"[配比] {args.total} 词/日：" +
          "、".join(f"{n} {int(q*100)}%" for n, q in RATIO) +
          f" -> " + "、".join(f"{n}{quota[n]}词" for n in CATEGORIES), file=sys.stderr)

    rows = []
    this_day_kws = []
    rotated_cats = []
    for cat in CATEGORIES:
        sub = df[df["类别"] == cat]
        q = quota[cat]
        used_cat = used_by_cat.get(cat, [])
        chosen, rotated = pick_for_category(sub, used_cat, q)
        intent = INTENT[cat]
        for kw in chosen:
            rows.append((cat, kw, intent))
            this_day_kws.append(kw)
        # 更新该组历史：轮转时重置为当天新取的词（保证次日继续往下排不重复），
        # 未轮转则并入历史
        if rotated:
            used_by_cat[cat] = sorted(set(chosen))
            rotated_cats.append(cat)
        else:
            used_by_cat[cat] = sorted(set(used_cat) | set(chosen))

    # 输出
    title = f"WPS SEO 关键词每日规划 · 第 {day} 天"
    sep = "=" * 60
    lines = [sep, title, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
             f"词量：{args.total} 词（品牌{quota['品牌词']} / 竞品{quota['竞品词']} / 通用{quota['通用词']} / 功能{quota['功能词']}）",
             sep, "", "| 策略组别 | 今日关键词 | 搜索意图 |",
             "| --- | --- | --- |"]
    for g, k, i in rows:
        lines.append(f"| {g} | {k} | {i} |")

    # 覆盖统计
    used_total = len(set().union(*[set(v) for v in used_by_cat.values()] or [set()]))
    remaining = len(df) - len(set(this_day_kws) | set().union(*[set(v) for v in used_by_cat.values()] or [set()]))
    lines.append("")
    lines.append("---")
    lines.append("**覆盖说明**")
    lines.append(f"- 今日新增：{len(this_day_kws)} 词（累计已用 {used_total} 词，"
                 f"词库剩余可覆盖 {max(remaining, 0)} 词）")
    if rotated_cats:
        lines.append(f"- 以下组别未用词已用尽，已自动轮转重头取词：{('、'.join(rotated_cats))}")

    text = "\n".join(lines)
    print(text)

    # 更新状态（无论是否 --out 都持久化去重记录）
    save_state(state_path, {"last_day": day, "used_by_cat": used_by_cat,
                            "last_output": f"第{day}天", "updated_at": None})
    print(f"\n[状态] 去重记录已更新：{state_path}", file=sys.stderr)

    # 可选导出
    if args.out:
        out = os.path.abspath(args.out)
        if out.lower().endswith(".xlsx"):
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"第{day}天规划"
                ws.append(["策略组别", "今日关键词", "搜索意图"])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DDEBF7")
                for g, k, i in rows:
                    ws.append([g, k, i])
                wb.save(out)
            except ImportError:
                sys.exit("[错误] 导出 xlsx 需要 openpyxl，请先安装")
        else:
            with open(out, "w", encoding="utf-8") as f:
                f.write(text + "\n")
        print(f"[导出] {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
