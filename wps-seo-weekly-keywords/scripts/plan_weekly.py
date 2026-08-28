#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WPS SEO 组别关键词周度规划脚本
================================
从《WPS_SEO完整关键词库》（wps_final_kw.csv）读取关键词，
按 10 个组别（意图→策略类型）每组挑取价值最高的若干关键词，
输出「策略组别 / 本周关键词 / 搜索意图」三列表，并自动去重：
每次调用给出的关键词与历史已用词不重复。

用法示例：
    python plan_weekly.py
    python plan_weekly.py --week 1
    python plan_weekly.py --per-group 5
    python plan_weekly.py --out /path/to/第2周规划.md
    python plan_weekly.py --csv /custom/path/wps_final_kw.csv --state /custom/path/state.json
    python plan_weekly.py --reset   # 清空去重历史，从第 1 周重新开始
"""
import argparse
import glob
import json
import os
import sys
from datetime import datetime

# ---------------------------------------------------------------------------
# 10 组别定义：组名 -> (筛选谓词, 意图标签)
# 筛选谓词接收 DataFrame，返回布尔 Series。
# ---------------------------------------------------------------------------
def _df(df):  # noqa: 仅为占位，实际谓词直接用闭包
    return df

GROUPS = [
    ("下载安装", "下载落地页",
     lambda d: d["一级分类"] == "下载安装",
     "下载WPS各端安装包，落地页需官方下载入口承接"),
    ("价格购买", "价格/会员页",
     lambda d: d["一级分类"] == "价格与会员",
     "会员价格/免费权益/版本对比，适合价格与权益展示页"),
    ("对比选择", "竞品对比页",
     lambda d: d["一级分类"] == "竞品对比与选型",
     "对比选型阶段，适合客观对比页与FAQ"),
    ("模板获取", "模板聚合页",
     lambda d: d["一级分类"] == "模板业务",
     "简历/汇报/合同/表格等模板需求，适合模板库聚合页"),
    ("格式转换", "转换工具页",
     lambda d: d["一级分类"] == "格式转换与文档处理",
     "PDF/Word/PPT/图片互转与处理，适合工具页与教程"),
    ("故障解决", "FAQ解决页",
     lambda d: (d["一级分类"] == "组件功能与教程")
               & (d["二级主题"] == "WPS用户问题/故障解决"),
     "WPS使用故障/报错/文件丢失恢复问题，适合FAQ解决页与GEO答案资产"),
    ("教程操作", "教程内容页",
     lambda d: (d["一级分类"] == "组件功能与教程")
               & (d["二级主题"] != "WPS用户问题/故障解决"),
     "WPS表格/文字/演示/PDF功能与操作，适合功能教程页"),
    ("功能认知", "功能科普页",
     lambda d: d["一级分类"] == "功能教程与知识",
     "办公知识与操作问答，适合教程内容与GEO答案资产"),
    ("AI认知", "AI专题页",
     lambda d: d["一级分类"] == "AI办公",
     "AI生成PPT/写作/表格，适合WPS AI专题页与GEO抢占"),
    ("品牌直达", "品牌落地页",
     lambda d: d["一级分类"] == "品牌与官方入口",
     "品牌/官网/正版入口，路径短、转化快，重承接与信任"),
]

# 价值列优先级：越靠前越优先作为排序依据
VALUE_COLS = ["GEO机会值", "第三方流量指数", "百度统计PV", "社媒热度值"]


# ---------------------------------------------------------------------------
# 路径定位
# ---------------------------------------------------------------------------
def find_workspace_csv():
    """按优先级定位词库 CSV：显式参数 > 当前/最近工作区 > 技能内置副本"""
    candidates = []
    ws = os.environ.get("WORKSPACE_DIR")
    if ws:
        candidates.append(os.path.join(ws, "wps_final_kw.csv"))
    # 常见工作区根目录
    home = os.path.expanduser("~")
    for base in [os.path.join(home, "Documents", "lingxi-claw")]:
        if os.path.isdir(base):
            candidates += sorted(glob.glob(os.path.join(base, "*", "wps_final_kw.csv")),
                                 key=os.path.getmtime, reverse=True)
    # 技能内置副本
    skill_data = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                              "assets", "data", "wps_final_kw.csv")
    candidates.append(skill_data)
    for c in candidates:
        if c and os.path.isfile(c):
            return c
    return None


def find_state_path(explicit=None):
    """状态文件：显式参数 > 最近工作区已有状态 > 最近工作区默认落点 > 技能目录 data/"""
    if explicit:
        return explicit
    home = os.path.expanduser("~")
    base = os.path.join(home, "Documents", "lingxi-claw")
    if os.path.isdir(base):
        found = sorted(glob.glob(os.path.join(base, "*", ".wps_seo_state.json")),
                       key=os.path.getmtime, reverse=True)
        if found:
            return found[0]
        # 未创建过：默认落到最近的工作区根目录，随工作区长期保留
        wss = sorted([d for d in glob.glob(os.path.join(base, "*")) if os.path.isdir(d)],
                     key=os.path.getmtime, reverse=True)
        if wss:
            return os.path.join(wss[0], ".wps_seo_state.json")
    ws = os.environ.get("WORKSPACE_DIR")
    if ws:
        return os.path.join(ws, ".wps_seo_state.json")
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.makedirs(os.path.join(skill_root, "data"), exist_ok=True)
    return os.path.join(skill_root, "data", "state.json")


# ---------------------------------------------------------------------------
# 状态读写
# ---------------------------------------------------------------------------
def load_state(path):
    if os.path.isfile(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"last_week": 0, "used": [], "updated_at": None}


def save_state(path, state):
    state["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def pick_for_group(sub, used, per_group):
    """从子集 sub 中，剔除 used 后按价值降序取 per_group 个。
    返回 (chosen, exhausted)。exhausted=True 表示组内未用词已用尽，触发轮转。"""
    cand = sub[~sub["关键词"].isin(used)].copy()
    exhausted = len(cand) < per_group
    if cand.empty:
        # 整组用尽：重置该组历史，从头再取
        cand = sub.copy()
    # 价值排序：取第一非空价值列作为分数
    def _score(r):
        for c in VALUE_COLS:
            v = r[c]
            if v is not None and not (isinstance(v, float) and v != v):  # 非 NaN
                try:
                    return float(v)
                except (TypeError, ValueError):
                    continue
        return 0.0

    cand["_score"] = cand.apply(_score, axis=1)
    cand = cand.sort_values("_score", ascending=False)
    chosen = cand.head(per_group)["关键词"].tolist()
    return chosen, exhausted


def main():
    ap = argparse.ArgumentParser(description="WPS SEO 组别关键词周度规划")
    ap.add_argument("--csv", help="词库 CSV 路径（默认自动定位）")
    ap.add_argument("--state", help="状态 JSON 路径（默认自动定位）")
    ap.add_argument("--week", type=int, help="指定周次（默认自动递增）")
    ap.add_argument("--per-group", type=int, default=5, help="每组关键词数（默认 5）")
    ap.add_argument("--out", help="输出文件路径（.md 或 .xlsx）")
    ap.add_argument("--reset", action="store_true", help="清空去重历史，从第 1 周开始")
    args = ap.parse_args()

    import pandas as pd

    csv_path = args.csv or find_workspace_csv()
    if not csv_path or not os.path.isfile(csv_path):
        sys.exit("[错误] 找不到词库文件 wps_final_kw.csv，请用 --csv 指定路径")
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    df = df.fillna({c: None for c in VALUE_COLS})
    print(f"[数据] 词库：{os.path.basename(csv_path)}（{len(df)} 词）", file=sys.stderr)

    state_path = find_state_path(args.state)
    if args.reset:
        state = {"last_week": 0, "used": [], "updated_at": None}
    else:
        state = load_state(state_path)
    used = set(state.get("used", []))
    week = args.week if args.week else (state.get("last_week", 0) + 1)

    rows = []
    exhausted_groups = []
    this_week_kws = []
    for gname, strategy, pred, intent in GROUPS:
        sub = df[pred(df)]
        chosen, exhausted = pick_for_group(sub, used, args.per_group)
        for kw in chosen:
            rows.append((f"{gname} · {strategy}", kw, intent))
            this_week_kws.append(kw)
        if exhausted:
            exhausted_groups.append(gname)

    # 输出
    title = f"WPS SEO 关键词周度规划 · 第 {week} 周"
    sep = "=" * 60
    lines = [sep, title, f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
             f"覆盖组别：{len(GROUPS)} 组 × {args.per_group} 词 = {len(this_week_kws)} 词",
             sep, "", "| 策略组别 | 本周关键词 | 搜索意图 |",
             "| --- | --- | --- |"]
    for g, k, i in rows:
        lines.append(f"| {g} | {k} | {i} |")

    # 覆盖统计
    remaining = len(df) - len(set(used) | set(this_week_kws))
    lines.append("")
    lines.append("---")
    lines.append("**覆盖说明**")
    lines.append(f"- 本周新增：{len(this_week_kws)} 词（累计已用 {len(set(used) | set(this_week_kws))} 词，"
                 f"词库剩余可覆盖 {max(remaining, 0)} 词）")
    if exhausted_groups:
        lines.append(f"- 以下组别未用词已用尽，已自动轮转重头取词：{('、'.join(exhausted_groups))}")

    text = "\n".join(lines)
    print(text)

    # 更新状态（无论是否 --out 都持久化去重记录）
    new_used = sorted(set(used) | set(this_week_kws))
    save_state(state_path, {"last_week": week, "used": new_used,
                            "last_output": f"第{week}周", "updated_at": None})
    print(f"\n[状态] 去重记录已更新：{state_path}", file=sys.stderr)

    # 可选导出
    if args.out:
        out = os.path.abspath(args.out)
        if out.lower().endswith(".xlsx"):
            try:
                import openpyxl  # noqa
                from openpyxl.styles import Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"第{week}周规划"
                ws.append(["策略组别", "本周关键词", "搜索意图"])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DDEBF7")
                for g, k, i in rows:
                    ws.append([g, k, i])
                wb.save(out)
            except ImportError:
                sys.exit("[错误] 导出 xlsx 需要 openpyxl，请先安装")
        else:
            with open(out, "w", encoding="utf-8") as f:
                f.write(text + "\n")
        print(f"[导出] {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
